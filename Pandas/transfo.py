import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fastapi import FastAPI, UploadFile, File

# app=FastAPI()

# @app.post("/upload")
# def upload_file(file: UploadFile = File(...)):
#     data = pd.read_excel(file.file, usecols="A:E,G:I,V:X,Z")
#     output_path = "crm_ready.xlsx"
#     data.to_excel(output_path, index=False)
#     return FileResponse(
#         output_path,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         filename="crm_ready.xlsx"
#     )

red_fill = PatternFill(
    start_color="FF0000",
    end_color="FF0000",
    fill_type="solid"
)

start_row = 2
valid_domains = (
    "@gmail.com",
    "@yahoo.com",
    "@outlook.com",
    "@hotmail.com",
    "@live.com",
    "@icloud.com", 
    "@msn.com",
    "@laposte.net",
    "@orange.fr",
    "@sfr.fr",
    "@free.fr",
    "@wanadoo.fr",
    "@aol.com",
    "@protonmail.com",
    "@gmx.com",
    "@gmx.fr",
    "@mail.com"
)


if __name__== "__main__":
    data = pd.read_excel("exemple.xlsx", usecols="A:E,G,V:X,Z", dtype={"Téléphone": str})
    
    wb = load_workbook("Matrice import CRM.xlsx")
    ws = wb["Template"]
    
    # parent=data[data.profiles == "Parent"].reset_index(drop=True)
    # student=data[data.profiles.isin(["Lycéen", "Collégien", "Etudiant"])].reset_index(drop=True)
    data["profiles"] = data["profiles"].replace({
        "Lycéen": "Elève, étudiant",
        "Collégien": "Elève, étudiant",
        "Etudiant": "Elève, étudiant"
    })
    # export=pd.concat([student,parent])
    # print(data)
    
    for i, row in data.iterrows():
        ws.cell(row=start_row + i, column=3, value=row["profiles"])
        ws.cell(row=start_row + i, column=5, value=row["Nom"])
        ws.cell(row=start_row + i, column=6, value=row["Prénom"])
        
        
        if not str(row["Email"]).endswith(valid_domains):
            ws.cell(row=start_row + i, column=7, value=row["Email"]).fill = red_fill
        else:
            ws.cell(row=start_row + i, column=7, value=row["Email"])
        
        
        if not str(row["Téléphone"]).startswith("+"):
            ws.cell(row=start_row + i, column=8, value=row["Téléphone"]).fill = red_fill
        else:
            ws.cell(row=start_row + i, column=8, value=row["Téléphone"])
        
        ws.cell(row=start_row + i, column=12, value=row["zipcode"])
        ws.cell(row=start_row + i, column=15, value=row["Actuellement, l’étudiant est en :"])
        ws.cell(row=start_row + i, column=22, value=row["Choix de campus :"])
        ws.cell(row=start_row + i, column=23, value=row["Souhaits de formations :"])
        # ws.cell(row=start_row + i, column=8, value=row["Intérêts pour les produits"])

    wb.save("export_crm2.xlsx")
    # output="export_crm.xlsx"
    # data.to_excel(output, index=False)
    
    
    # print(data)
    # export=parent

    # print(student)
    # print(parent)
    # student.to_excel("output.xls")
    
    # app=FastAPI()