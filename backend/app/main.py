from fastapi import FastAPI, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from app.api.routes import router
from app.database import engine, Base
from app.models import user
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

# Créer les tables dans la base de données
Base.metadata.create_all(bind=engine)

app = FastAPI(title="SA Plateforme API", version="1.0.0")

# Configuration CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

#Couleur rouge pour les données erronées
red_fill = PatternFill(
    start_color="FF0000",
    end_color="FF0000",
    fill_type="solid"
)

start_row = 2

# Domain possible pour les adresses mails
valid_domains = (
    "@gmail.com",
    "@yahoo.com",
    "@outlook.com",
    "@hotmail.com",
    "@live.com",
    "@icloud.com", 
    "@msn.com",
    "@laposte.net",
    "@orange.fr",
    "@sfr.fr",
    "@free.fr",
    "@wanadoo.fr",
    "@aol.com",
    "@protonmail.com",
    "@gmx.com",
    "@gmx.fr",
    "@mail.com"
)

# Inclusion des routes
app.include_router(router, prefix="/api")

@app.get("/")
async def root():
    return {"message": "Bienvenue sur l'API SA Plateforme", "docs": "/docs"}

# Fonction pour traiter les fichiers
@app.post("/upload")
def upload_file(file: UploadFile = File(...)):
    data = pd.read_excel(file.file, usecols="A:E,G:I,V:X,Z", dtype={"Téléphone": str})
    
    # Template du fichier CRM
    wb = load_workbook("Matrice import CRM.xlsx")
    ws = wb["Template"]
    
    #  Transformation de certains profils en "Elève, étudiant"
    data["profiles"] = data["profiles"].replace({
        "Lycéen": "Elève, étudiant",
        "Collégien": "Elève, étudiant",
        "Etudiant": "Elève, étudiant"
    })
    
    # Ajout des données dans le fichier CRM "Template"
    for i, row in data.iterrows():
        ws.cell(row=start_row + i, column=3, value=row["profiles"])
        ws.cell(row=start_row + i, column=5, value=row["Nom"])
        ws.cell(row=start_row + i, column=6, value=row["Prénom"])
        
        # Vérification des emails
        if not str(row["Email"]).endswith(valid_domains):
            ws.cell(row=start_row + i, column=7, value=row["Email"]).fill = red_fill
        else:
            ws.cell(row=start_row + i, column=7, value=row["Email"])
        
        # Vérification des numéros de téléphones
        if not str(row["Téléphone"]).startswith("+") or len(str(row["Téléphone"])) != 12:
            ws.cell(row=start_row + i, column=8, value=row["Téléphone"]).fill = red_fill
        else:
            ws.cell(row=start_row + i, column=8, value=row["Téléphone"])
        
        ws.cell(row=start_row + i, column=12, value=row["zipcode"])
        ws.cell(row=start_row + i, column=15, value=row["Actuellement, l’étudiant est en :"])
        ws.cell(row=start_row + i, column=22, value=row["Choix de campus :"])
        ws.cell(row=start_row + i, column=23, value=row["Souhaits de formations :"])
    
    # Export du fichier créé
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=crm_ready.xlsx"
        }
    )